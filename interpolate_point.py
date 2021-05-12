import sys
sys.path.append("lib")
from Interpolate import interpolate,calculate_wind_direction
import Readtime
from wrf import getvar,to_np
from openpyxl import Workbook
import numpy as np
import netCDF4 as nc

#下面修改路径
path='D:\wrf_simulation\\2meic\wrfout_d03_2016-07-21_12_2meic'

#下面修改站点的纬度，经度，气象站点名字。气象站点名字可以带中文，主要就是为了excel表格的sheet书写
point_list=[(31.1,121.37,"58361闵行"),(31.39692,121.45454,"58362宝山"),(31.37,121.25,"58365嘉定"),(31.67,121.50,"58366崇明"),
            (31.05,121.7833,"58369南汇"),(31.13,121.12,"58461青浦"),(30.88,121.50,"58463奉贤")]
ncfile=nc.Dataset(path)
wb=Workbook()
for i in point_list:
    wb.create_sheet(str(i[2]))
    ws = wb[str(i[2])]
    ws.cell(1,1,"时间")
    ws.cell(1,2,"2m温度")
    ws.cell(1,3,"2m湿度")
    ws.cell(1,4,"10m风向")
    ws.cell(1,5,"10m风速")
    ws.cell(1,6,"u")
    ws.cell(1,7,"v")

timelist=Readtime.get_ncfile_time(ncfile)
for i in range(0,Readtime.get_ncfile_alltime(ncfile),3):
    t2=to_np(getvar(ncfile,"T2",timeidx=i))-273.15
    rh2=to_np(getvar(ncfile,'rh2',timeidx=i))
    u10 = to_np(getvar(ncfile, "U10", timeidx=i))
    v10 = to_np(getvar(ncfile,"V10",timeidx=i))
    for j in range(len(point_list)):
        float_t2 = interpolate(ncfile, t2, point_list[j][0], point_list[j][1], opt=0)
        float_rh2 = interpolate(ncfile, rh2, point_list[j][0], point_list[j][1], opt=0)
        float_u10 = interpolate(ncfile, u10, point_list[j][0], point_list[j][1], opt=0)
        float_v10 = interpolate(ncfile, v10, point_list[j][0], point_list[j][1], opt=0)
        print("u10,v10:" + str(float_u10) + "," + str(float_v10))
        float_ws = np.sqrt(float_v10**2+float_u10**2)
        float_wdir = calculate_wind_direction(float_u10,float_v10)
        print(timelist[i])
        print("u10,v10:"+str(float_u10)+","+str(float_v10)+","+str(float_ws)+","+str(float_wdir))
        print(i/3)
        worksheet=wb[str(point_list[j][2])]
        worksheet.cell(i/3+2,1,timelist[i])
        worksheet.cell(i/3+2,2,float_t2)
        worksheet.cell(i/3+2,3,float_rh2)
        worksheet.cell(i/3+2,4,float_wdir)
        worksheet.cell(i/3+2,5,float_ws)
        worksheet.cell(i/3+2,6,float_u10)
        worksheet.cell(i/3+2,7,float_v10)
wb.save("气象信息.xlsx")




