#该脚本用于上海市地面气象资料的合并与计算
from datetime import datetime
from openpyxl import Workbook,load_workbook
import os
import numpy as np

def creat_xlsx(filename,station_num,lat,lon,height):
    wb=Workbook()
    ws=wb.active
    ws.cell(1,1,'年')
    ws.cell(1,2,'月')
    ws.cell(1,3,'日')
    ws.cell(1,4,'时')
    ws.cell(1,5,'海平面气压')
    ws.cell(1,6,'温度')
    ws.cell(1,7,'露点温度')
    ws.cell(1,8,'相对湿度')
    ws.cell(1,9,'风向')
    ws.cell(1,10,'风速')
    ws.cell(1,11,'观测前1小时降水')
    ws.cell(1,12,'观测前6小时降水')
    ws.cell(1,13,'观测前24小时降水')
    wb.create_sheet("站点信息")
    ws2=wb["站点信息"]
    ws2.cell(1,1,"站点名称:")
    ws2.cell(1,2,station_num)
    ws2.cell(2,1,"经度：")
    ws2.cell(2,2,lon)
    ws2.cell(3,1,"纬度：")
    ws2.cell(3,2,lat)
    ws2.cell(4,1,"站点高度：")
    ws2.cell(4,2,height)
    wb.save(filename)

def get_xlsx_time(filename,datetime_):
    exist_file=load_workbook(filename)
    exist_ws=exist_file['Sheet']
    n=2
    if exist_ws.max_row==1:
        return 2
    else:
        last_n=exist_ws.max_row
        last_datetime=datetime(int(exist_ws.cell(last_n,1).value),int(exist_ws.cell(last_n,2).value),
                               int(exist_ws.cell(last_n,3).value),int(exist_ws.cell(last_n,4).value))
        if datetime_>last_datetime:
            return last_n+1
        if datetime_<last_datetime:
            while True:
                exist_datetime = datetime(int(exist_ws.cell(n, 1).value), int(exist_ws.cell(n, 2).value),
                                         int(exist_ws.cell(n, 3).value), int(exist_ws.cell(n, 4).value))
                if datetime_>exist_datetime:
                    return n+1
                n+=1

def write_xlsx(filename,row,strlist,date):
    wb=load_workbook(filename)
    ws=wb['Sheet']
    strlist=np.array(strlist)
    floatlist=strlist.astype(float)
    a,b=17.27,237.7
    t,td=floatlist[8],floatlist[9]
    c=a*t/(b+t)
    rh=np.e**((td*a-td*c-b*c)/(b+td))*100
    ws.cell(row,1,date[0])
    ws.cell(row,2,date[1])
    ws.cell(row,3,date[2])
    ws.cell(row,4,date[3])
    ws.cell(row,5,floatlist[7])
    ws.cell(row,6,floatlist[8])
    ws.cell(row,7,floatlist[9])
    ws.cell(row,8,rh)
    ws.cell(row,9,floatlist[10])
    ws.cell(row,10,floatlist[11])
    ws.cell(row,11,floatlist[12])
    ws.cell(row,12,floatlist[13])
    ws.cell(row,13,floatlist[14])
    wb.save(filename)




path = "E:\数据及ncl脚本\地面资料" #文件夹目录
save_path="E:\数据及ncl脚本"
files= os.listdir(path) #得到文件夹下的所有文件名称
s = []
for file in files: #遍历文件夹
    f = open(path+"/"+file,encoding='UTF-8') #打开文件
    n=True
    for line in f: #遍历文件，一行行遍历，读取文本
        if n==True:
            print(str(file)+"第一行数据")
            n=False
            strlist = line.split()
            datetime_list=np.array(strlist)
            datetime_list_ymdm=[int(datetime_list[1]),int(datetime_list[2]),int(datetime_list[3]),int(datetime_list[4])]
            datetime_=datetime(int(datetime_list[1]),int(datetime_list[2]),int(datetime_list[3]),int(datetime_list[4]))
        else:
            strlist=line.split()
            strlist=np.array(strlist)
            if os.path.exists(save_path+"/"+strlist[0]+".xlsx"):
                print(strlist[0]+"文件已存在,将追加读写")
            else:
                creat_xlsx(save_path+"/"+strlist[0]+".xlsx",strlist[0],strlist[1],strlist[2],strlist[3])
                print(strlist[0] + "文件不存在，将创建")
            row=get_xlsx_time(save_path+"/"+strlist[0]+".xlsx",datetime_)
            write_xlsx(save_path+"/"+strlist[0]+".xlsx",row,strlist,datetime_list_ymdm)
            print(strlist[0],datetime_,"   ",row)
